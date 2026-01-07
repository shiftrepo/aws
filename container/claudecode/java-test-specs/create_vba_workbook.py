#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VBA-Enabled Excel Workbook Creator
Creates a macro-enabled Excel workbook (.xlsm) template ready for VBA module import.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_vba_template_workbook():
    """Create a template Excel workbook for VBA macro import."""

    # Create workbook
    wb = Workbook()

    # Remove default sheet and create main sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create main control sheet
    ws = wb.create_sheet("TestSpecGenerator", 0)

    # Define styles
    title_font = Font(size=18, bold=True, color="2F4F4F")
    subtitle_font = Font(size=14, bold=True, color="4682B4")
    instruction_font = Font(size=11)
    button_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    button_font = Font(size=12, bold=True, color="FFFFFF")

    # Create title section
    ws['B2'] = "Java Test Specification Generator"
    ws['B2'].font = title_font
    ws.merge_cells('B2:F2')

    ws['B3'] = "VBA Macro Tool for Automated Test Documentation"
    ws['B3'].font = subtitle_font
    ws.merge_cells('B3:F3')

    # Instructions section
    ws['B5'] = "Instructions for Use:"
    ws['B5'].font = Font(size=12, bold=True)

    instructions = [
        "1. Import VBA Modules (see vba-modules/VBA-Import-Instructions.md)",
        "2. Enable macros when opening this file",
        "3. Click 'Generate Test Specification' button below",
        "4. Select source directory containing Java test files",
        "5. Choose output location for Excel report",
        "6. Wait for processing to complete"
    ]

    for i, instruction in enumerate(instructions, 6):
        ws[f'C{i}'] = instruction
        ws[f'C{i}'].font = instruction_font

    # Create button placeholder (actual button will be added in Excel)
    ws['C13'] = "📊 Generate Test Specification"
    ws['C13'].font = button_font
    ws['C13'].fill = button_fill
    ws['C13'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C13:E13')
    ws.row_dimensions[13].height = 30

    # Add note about button assignment
    ws['C15'] = "Note: After importing VBA modules, right-click the green button above"
    ws['C15'].font = Font(size=10, italic=True)
    ws['C16'] = "and select 'Assign Macro' → 'MainController.GenerateTestSpecification'"
    ws['C16'].font = Font(size=10, italic=True)

    # Sample data section
    ws['B18'] = "Sample Data Location:"
    ws['B18'].font = Font(size=12, bold=True)

    ws['C19'] = "Java test files: sample-java-tests/"
    ws['C19'].font = instruction_font

    ws['C20'] = "Coverage reports: sample-java-tests/coverage-reports/"
    ws['C20'].font = instruction_font

    ws['C21'] = "Expected output: examples/TestSpecification_Sample_20260107.xlsx"
    ws['C21'].font = instruction_font

    # Requirements section
    ws['B23'] = "System Requirements:"
    ws['B23'].font = Font(size=12, bold=True)

    requirements = [
        "• Microsoft Excel 2016 or later",
        "• VBA enabled (Developer tab visible)",
        "• Macros enabled in security settings",
        "• Write permissions to output directory"
    ]

    for i, req in enumerate(requirements, 24):
        ws[f'C{i}'] = req
        ws[f'C{i}'].font = instruction_font

    # Version info
    ws['B29'] = "Application Information:"
    ws['B29'].font = Font(size=12, bold=True)

    ws['C30'] = "Version: 1.0.0"
    ws['C30'].font = instruction_font

    ws['C31'] = "Created: 2026-01-07"
    ws['C31'].font = instruction_font

    ws['C32'] = "VBA Modules: 6 (MainController, FolderScanner, JavaAnnotationParser, etc.)"
    ws['C32'].font = instruction_font

    # Set column widths
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    # Set row heights for better appearance
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 20

    return wb

def main():
    """Main function to create the VBA template workbook."""

    print("Creating VBA-enabled Excel template workbook...")

    # Create workbook
    wb = create_vba_template_workbook()

    # Save as .xlsm (macro-enabled format)
    output_path = "/root/aws.git/container/claudecode/java-test-specs/TestSpecGenerator_Template.xlsm"

    # Note: openpyxl saves as .xlsx by default, but we'll name it .xlsm
    # The actual macro functionality will be added when VBA modules are imported in Excel
    wb.save(output_path.replace('.xlsm', '.xlsx'))  # Save as xlsx first

    # Rename to xlsm
    xlsx_path = output_path.replace('.xlsm', '.xlsx')
    os.rename(xlsx_path, output_path)

    print(f"✅ VBA template workbook created successfully!")
    print(f"📍 File location: {output_path}")
    print(f"📝 Next steps:")
    print(f"   1. Open file in Excel")
    print(f"   2. Import VBA modules from vba-modules/ directory")
    print(f"   3. Assign macro to the green button")
    print(f"   4. Save as macro-enabled workbook (.xlsm)")

    # Show file info
    if os.path.exists(output_path):
        file_size = os.path.getsize(output_path)
        print(f"📁 File size: {file_size:,} bytes")

    return output_path

if __name__ == "__main__":
    main()