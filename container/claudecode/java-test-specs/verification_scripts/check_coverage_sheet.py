import sys
import warnings
warnings.filterwarnings('ignore')

try:
    import openpyxl
    from openpyxl import load_workbook
    
    # Excelファイルを読み込み
    wb = load_workbook('improved_test_specification.xlsx', read_only=True)
    
    # Coverageシートを取得
    if 'Coverage' in wb.sheetnames:
        sheet = wb['Coverage']
        
        print("=== Improved Coverage Sheet Content (First 30 rows) ===\n")
        
        # 各行をテキストとして表示（最大30行まで）
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            if row_count >= 30:  # 最初の30行のみ表示
                print("\n... (以降省略)")
                # 全行数をカウント
                total_rows = sum(1 for _ in sheet.iter_rows())
                print(f"合計行数: {total_rows}")
                break
            
            # Noneを空文字列に変換して表示
            row_values = [str(cell) if cell is not None else "" for cell in row]
            # 特定の列のみ表示（No, Package, Class, Method, Source File）
            if row_count == 0:
                print(" | ".join([row_values[0], row_values[1], row_values[2], row_values[3], row_values[4]]))
            else:
                print(f"{row_values[0]:3} | {row_values[1]:30} | {row_values[2]:30} | {row_values[3]:20} | {row_values[4]:30}")
            row_count += 1
    else:
        print("Coverageシートが見つかりません")
        print("利用可能なシート:", wb.sheetnames)
    
    wb.close()
    
except ImportError:
    print("openpyxlがインストールされていません。")
except FileNotFoundError:
    print("Excelファイルが見つかりません")
except Exception as e:
    print(f"エラー: {e}")
