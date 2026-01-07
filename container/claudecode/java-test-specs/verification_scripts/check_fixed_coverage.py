import warnings
warnings.filterwarnings('ignore')

import openpyxl
from openpyxl import load_workbook

# Check the fixed Excel file
wb = load_workbook('fixed_coverage_test_specification.xlsx', read_only=True)

# Check Coverage sheet
if 'Coverage' in wb.sheetnames:
    sheet = wb['Coverage']

    print("=== Fixed Coverage Sheet Check (First 20 rows) ===\n")

    # Read first 20 rows
    row_count = 0
    for row in sheet.iter_rows(values_only=True):
        if row_count >= 20:
            break

        # Display all columns
        if row_count == 0:
            print(f"Headers: {row}")
        else:
            print(f"Row {row_count}: {row}")
        row_count += 1

    # Count total rows
    total_rows = sum(1 for _ in sheet.iter_rows()) - 1  # Exclude header
    print(f"\nTotal data rows: {total_rows}")

wb.close()