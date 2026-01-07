#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Java Test Specification Generator - Sample Excel Report Creator
Creates a sample Excel report demonstrating the output from the VBA tool.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_test_specification_excel():
    """Create a comprehensive test specification Excel file with sample data."""

    # Create workbook and remove default sheet
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Define colors
    header_blue = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    light_blue = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    header_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    header_yellow = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    black_font = Font(color="000000", bold=True)

    # Define borders
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Create Test Details sheet
    create_test_details_sheet(wb, header_blue, white_font, thin_border)

    # Create Summary sheet
    create_summary_sheet(wb, light_blue, black_font, thin_border)

    # Create Coverage sheet
    create_coverage_sheet(wb, header_green, black_font, thin_border)

    # Create Configuration sheet
    create_configuration_sheet(wb, header_yellow, black_font, thin_border)

    return wb

def create_test_details_sheet(wb, header_fill, header_font, border):
    """Create the Test Details sheet with comprehensive test case information."""

    ws = wb.create_sheet("Test Details", 0)

    # Title section
    ws['A1'] = "Test Specification Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:O1')

    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(bold=True)

    ws['A3'] = "Source Directory: /root/aws.git/container/claudecode/java-test-specs/sample-java-tests"
    ws['A3'].font = Font(bold=True)

    # Headers
    headers = [
        "File Path", "Test Module", "Test Case", "Baseline Version", "Test Overview",
        "Test Purpose", "Test Process", "Test Results", "Creator", "Created Date",
        "Modifier", "Modified Date", "Coverage %", "Branches Covered", "Branches Total"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True)

    # Sample data based on actual Java test files
    test_data = [
        [
            "BasicCalculatorTest.java", "CalculatorModule", "BasicArithmeticOperations", "1.0.0",
            "Verify basic calculator operations with conditional logic for C1 coverage",
            "Ensure proper handling of different numeric input types and edge cases",
            "Execute tests with various parameters to achieve condition/decision coverage",
            "All conditions should pass validation checks with proper branching",
            "DeveloperName", "2026-01-07", "ReviewerName", "2026-01-07", "94.7%", 36, 38
        ],
        [
            "BasicCalculatorTest.java", "CalculatorModule", "ConditionalAdditionTest", "1.0.0",
            "Test addition with conditional branching based on input values",
            "Demonstrate C1 coverage through parameter-based conditional logic",
            "Test positive, negative, and zero values to cover all branches",
            "Each condition branch should execute and pass assertions",
            "DeveloperName", "2026-01-07", "ReviewerName", "2026-01-07", "100.0%", 8, 8
        ],
        [
            "BasicCalculatorTest.java", "CalculatorModule", "MultiplicationBranching", "1.0.0",
            "Test multiplication with complex conditional logic",
            "Achieve comprehensive C1 coverage with nested conditions",
            "Test various numeric ranges with multiple decision points",
            "All conditional paths should be exercised and validated",
            "DeveloperName", "2026-01-07", "ReviewerName", "2026-01-07", "87.5%", 14, 16
        ],
        [
            "BasicCalculatorTest.java", "CalculatorModule", "DivisionWithValidation", "1.0.0",
            "Test division operations with error handling and validation",
            "Demonstrate exception handling and boundary condition testing",
            "Test normal division and division by zero scenarios",
            "Proper results for valid operations, exceptions for invalid ones",
            "DeveloperName", "2026-01-07", "ReviewerName", "2026-01-07", "100.0%", 12, 12
        ],
        [
            "StringValidatorTest.java", "StringValidationModule", "StringValidationOperations", "1.0.0",
            "Validate string inputs with comprehensive conditional logic for C1 coverage",
            "Ensure proper handling of various string types including null, empty, and valid strings",
            "Execute validation tests with different string parameters to achieve full condition coverage",
            "All string validation conditions should be properly tested and validated",
            "DeveloperName", "2026-01-07", "ReviewerName", "2026-01-07", "94.5%", 104, 110
        ],
        [
            "StringValidatorTest.java", "StringValidationModule", "EmailValidationTest", "1.0.0",
            "Test email validation with multiple conditional branches",
            "Demonstrate C1 coverage through different email format validations",
            "Test valid emails, invalid formats, null, and empty strings",
            "Each validation condition should execute properly and return correct boolean result",
            "DeveloperName", "2026-01-07", "ReviewerName", "2026-01-07", "95.8%", 23, 24
        ],
        [
            "StringValidatorTest.java", "StringValidationModule", "PasswordStrengthTest", "1.0.0",
            "Test password strength validation with complex conditional logic",
            "Achieve comprehensive C1 coverage through password criteria checking",
            "Test passwords of various lengths and complexity levels",
            "Password strength levels should be correctly categorized based on criteria",
            "DeveloperName", "2026-01-07", "QAEngineer", "2026-01-07", "90.6%", 29, 32
        ],
        [
            "StringValidatorTest.java", "StringValidationModule", "UsernameValidation", "1.0.0",
            "Test username validation with character and length restrictions",
            "Demonstrate multiple decision points for username acceptance criteria",
            "Test usernames with various characters, lengths, and formats",
            "Username validation should properly handle all edge cases and valid formats",
            "DeveloperName", "2026-01-07", "QAEngineer", "2026-01-07", "95.0%", 19, 20
        ]
    ]

    # Add data rows
    for row_idx, row_data in enumerate(test_data, 6):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx in [5, 6, 7, 8]:  # Text wrap for description columns
                cell.alignment = Alignment(wrap_text=True)
            if col_idx == 13:  # Coverage percentage
                cell.number_format = '0.0%'
                cell.value = float(value.replace('%', '')) / 100

    # Set column widths
    column_widths = [40, 15, 20, 12, 30, 30, 35, 30, 12, 12, 12, 12, 10, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_summary_sheet(wb, header_fill, header_font, border):
    """Create the Summary sheet with aggregated statistics."""

    ws = wb.create_sheet("Summary", 1)

    # Title
    ws['A1'] = "Test Summary Report"
    ws['A1'].font = Font(size=16, bold=True)

    # Basic statistics
    ws['A3'] = "Total Test Files:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = 2

    ws['A4'] = "Total Test Cases:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = 8

    ws['A5'] = "Total Test Methods:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = 9

    # Coverage statistics
    ws['A7'] = "Coverage Statistics"
    ws['A7'].font = Font(size=14, bold=True)

    ws['A8'] = "Overall Branch Coverage:"
    ws['A8'].font = Font(bold=True)
    ws['B8'] = 0.946  # 94.6%
    ws['B8'].number_format = '0.0%'

    ws['A9'] = "Branches Covered:"
    ws['A9'].font = Font(bold=True)
    ws['B9'] = 140

    ws['A10'] = "Total Branches:"
    ws['A10'].font = Font(bold=True)
    ws['B10'] = 148

    ws['A11'] = "Missing Coverage:"
    ws['A11'].font = Font(bold=True)
    ws['B11'] = 8

    # Processing information
    ws['A13'] = "Processing Information"
    ws['A13'].font = Font(size=14, bold=True)

    ws['A14'] = "Processing Time:"
    ws['A14'].font = Font(bold=True)
    ws['B14'] = "00:00:15"

    ws['A15'] = "Coverage Reports Found:"
    ws['A15'].font = Font(bold=True)
    ws['B15'] = 2

    # Module breakdown
    ws['A17'] = "Module Coverage Breakdown"
    ws['A17'].font = Font(size=14, bold=True)

    # Headers for module breakdown
    module_headers = ["Module Name", "Test Cases", "Coverage %", "Status"]
    for col, header in enumerate(module_headers, 1):
        cell = ws.cell(row=18, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Module data
    module_data = [
        ["CalculatorModule", 4, 0.947, "Excellent"],
        ["StringValidationModule", 4, 0.945, "Excellent"]
    ]

    for row_idx, row_data in enumerate(module_data, 19):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 3:  # Coverage percentage
                cell.number_format = '0.0%'

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def create_coverage_sheet(wb, header_fill, header_font, border):
    """Create the Coverage sheet with detailed coverage analysis."""

    ws = wb.create_sheet("Coverage", 2)

    # Title
    ws['A1'] = "Coverage Analysis Report"
    ws['A1'].font = Font(size=16, bold=True)

    # Headers
    coverage_headers = [
        "File Path", "Method Name", "Instructions Covered", "Instructions Missed",
        "Branches Covered", "Branches Missed", "C1 Coverage %", "Coverage Status"
    ]

    for col, header in enumerate(coverage_headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Coverage data based on JaCoCo report
    coverage_data = [
        ["BasicCalculatorTest.java", "testConditionalCalculation", 45, 2, 8, 0, 1.00, "Excellent"],
        ["BasicCalculatorTest.java", "testMultiplicationBranching", 78, 5, 14, 2, 0.875, "Good"],
        ["BasicCalculatorTest.java", "testDivisionWithValidation", 52, 3, 12, 0, 1.00, "Excellent"],
        ["BasicCalculatorTest.java", "[Class Total]", 182, 10, 36, 2, 0.947, "Excellent"],
        ["StringValidatorTest.java", "testEmailValidation", 95, 8, 23, 1, 0.958, "Excellent"],
        ["StringValidatorTest.java", "testPasswordStrengthValidation", 142, 12, 29, 3, 0.906, "Excellent"],
        ["StringValidatorTest.java", "testUsernameValidation", 78, 5, 19, 1, 0.95, "Excellent"],
        ["StringValidatorTest.java", "[Class Total]", 358, 25, 104, 6, 0.945, "Excellent"]
    ]

    # Add coverage data
    for row_idx, row_data in enumerate(coverage_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 7:  # Coverage percentage
                cell.number_format = '0.0%'

    # Set column widths
    column_widths = [30, 25, 15, 15, 15, 15, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_configuration_sheet(wb, header_fill, header_font, border):
    """Create the Configuration sheet with processing metadata."""

    ws = wb.create_sheet("Configuration", 3)

    # Title
    ws['A1'] = "Processing Configuration"
    ws['A1'].font = Font(size=16, bold=True)

    # Headers
    config_headers = ["Configuration Parameter", "Value", "Description"]
    for col, header in enumerate(config_headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Configuration data
    config_data = [
        ["Source Directory", "/root/aws.git/container/claudecode/java-test-specs/sample-java-tests", "Directory containing Java test files"],
        ["Output File", "TestSpecification_20260107_120000.xlsx", "Generated Excel report location"],
        ["Processing Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "When the report was generated"],
        ["Files Processed", "2", "Total number of Java files analyzed"],
        ["Test Cases Found", "8", "Total number of test cases extracted"],
        ["Coverage Reports Found", "2", "Number of JaCoCo reports processed"],
        ["Processing Duration", "00:00:15", "Time taken to generate this report"],
        ["Application Version", "1.0.0", "Version of the VBA application used"],
        ["Overall C1 Coverage", "94.6%", "Combined branch coverage percentage"],
        ["Total Branches", "148", "Total number of decision branches found"]
    ]

    # Add configuration data
    for row_idx, row_data in enumerate(config_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 1:  # Parameter names in bold
                cell.font = Font(bold=True)

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 40

def main():
    """Main function to create and save the Excel report."""

    print("Creating Java Test Specification Excel Report...")

    # Create workbook
    wb = create_test_specification_excel()

    # Save the file
    output_path = "/root/aws.git/container/claudecode/java-test-specs/examples/TestSpecification_Sample_20260107.xlsx"

    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Save workbook
    wb.save(output_path)

    print(f"✅ Excel report created successfully!")
    print(f"📍 File location: {output_path}")
    print(f"📊 Report contains 4 sheets:")
    print(f"   - Test Details: Complete test case information")
    print(f"   - Summary: Aggregated statistics")
    print(f"   - Coverage: Detailed C1 coverage analysis")
    print(f"   - Configuration: Processing metadata")

    # Show file info
    if os.path.exists(output_path):
        file_size = os.path.getsize(output_path)
        print(f"📁 File size: {file_size:,} bytes")

    return output_path

if __name__ == "__main__":
    main()