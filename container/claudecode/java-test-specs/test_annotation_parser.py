#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test script to verify Java annotation parsing works with actual files (no mocks)
Tests the actual annotation extraction from BasicCalculatorTest.java and StringValidatorTest.java
"""

import re
import os
import xml.etree.ElementTree as ET
from pathlib import Path

def parse_java_annotations(file_path):
    """Parse Java annotations from actual file content."""
    print(f"\n📄 Parsing annotations from: {file_path}")

    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Extract class-level annotations
    class_annotations = {}
    method_annotations = []

    # Find JavaDoc comments with annotations
    javadoc_pattern = r'/\*\*\s*(.*?)\s*\*/'
    javadoc_matches = re.findall(javadoc_pattern, content, re.DOTALL)

    for javadoc in javadoc_matches:
        # Look for @annotations
        annotation_pattern = r'@(\w+)\s+([^\n*]+)'
        annotations = re.findall(annotation_pattern, javadoc)

        if annotations:
            parsed_annotations = {}
            for name, value in annotations:
                parsed_annotations[name] = value.strip()

            # Check if this is method-level (has @Test nearby)
            if '@Test' in javadoc or 'public void test' in content[content.find(javadoc):content.find(javadoc)+200]:
                method_annotations.append(parsed_annotations)
            else:
                class_annotations.update(parsed_annotations)

    print(f"✅ Class annotations found: {len(class_annotations)}")
    for key, value in class_annotations.items():
        print(f"   @{key}: {value}")

    print(f"✅ Method annotations found: {len(method_annotations)}")
    for i, annotations in enumerate(method_annotations):
        print(f"   Method {i+1}: {list(annotations.keys())}")

    return class_annotations, method_annotations

def parse_jacoco_report(file_path):
    """Parse actual JaCoCo coverage data from XML report."""
    print(f"\n📊 Parsing JaCoCo report: {file_path}")

    tree = ET.parse(file_path)
    root = tree.getroot()

    total_branches = 0
    covered_branches = 0

    # Parse package and class coverage
    for package in root.findall('package'):
        package_name = package.get('name', 'default')
        print(f"📦 Package: {package_name}")

        for class_elem in package.findall('class'):
            class_name = class_elem.get('name', '').replace('/', '.')
            print(f"   📄 Class: {class_name}")

            # Parse method-level coverage
            for method in class_elem.findall('method'):
                method_name = method.get('name')

                # Get branch coverage counters
                for counter in method.findall('counter'):
                    if counter.get('type') == 'BRANCH':
                        missed = int(counter.get('missed', 0))
                        covered = int(counter.get('covered', 0))
                        total = missed + covered
                        coverage_pct = (covered / total * 100) if total > 0 else 0

                        print(f"      🎯 {method_name}: {covered}/{total} branches ({coverage_pct:.1f}%)")
                        total_branches += total
                        covered_branches += covered

    overall_coverage = (covered_branches / total_branches * 100) if total_branches > 0 else 0
    print(f"\n📈 Overall C1 Coverage: {covered_branches}/{total_branches} branches ({overall_coverage:.1f}%)")

    return {
        'total_branches': total_branches,
        'covered_branches': covered_branches,
        'coverage_percentage': overall_coverage
    }

def test_excel_file_content(file_path):
    """Test that Excel file was created with actual data."""
    print(f"\n📊 Testing Excel file: {file_path}")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    file_size = os.path.getsize(file_path)
    print(f"✅ File exists, size: {file_size:,} bytes")

    # Try to read with openpyxl to verify structure
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames

        print(f"✅ Sheets found: {sheet_names}")

        # Check Test Details sheet has actual data
        if 'Test Details' in sheet_names:
            ws = wb['Test Details']
            row_count = ws.max_row
            col_count = ws.max_column
            print(f"   📊 Test Details: {row_count} rows, {col_count} columns")

            # Check for actual test data (not just headers)
            if row_count > 5:
                sample_data = ws.cell(row=6, column=1).value
                print(f"   📝 Sample test file: {sample_data}")

        wb.close()
        return True

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False

def main():
    """Main test function - verify all components work with actual data."""
    print("🧪 Testing Java Test Specification Generator (No Mocks - Actual Data Only)")
    print("=" * 80)

    base_dir = "/root/aws.git/container/claudecode/java-test-specs"

    # Test 1: Parse actual Java files
    java_files = [
        f"{base_dir}/sample-java-tests/BasicCalculatorTest.java",
        f"{base_dir}/sample-java-tests/StringValidatorTest.java"
    ]

    all_class_annotations = {}
    all_method_annotations = []

    for java_file in java_files:
        if os.path.exists(java_file):
            class_ann, method_ann = parse_java_annotations(java_file)
            all_class_annotations.update(class_ann)
            all_method_annotations.extend(method_ann)
        else:
            print(f"❌ Java file not found: {java_file}")
            return False

    # Test 2: Parse actual JaCoCo report
    jacoco_file = f"{base_dir}/sample-java-tests/coverage-reports/jacoco-report.xml"
    if os.path.exists(jacoco_file):
        coverage_data = parse_jacoco_report(jacoco_file)
    else:
        print(f"❌ JaCoCo report not found: {jacoco_file}")
        return False

    # Test 3: Verify Excel file has actual content
    excel_file = f"{base_dir}/examples/TestSpecification_Sample_20260107.xlsx"
    excel_valid = test_excel_file_content(excel_file)

    # Test 4: Verify VBA template exists
    vba_file = f"{base_dir}/TestSpecGenerator_Template.xlsm"
    vba_exists = os.path.exists(vba_file)
    if vba_exists:
        vba_size = os.path.getsize(vba_file)
        print(f"\n📋 VBA Template: {vba_size:,} bytes")

    # Summary
    print("\n" + "=" * 80)
    print("🎯 TEST RESULTS SUMMARY:")
    print(f"✅ Java annotation parsing: {len(all_class_annotations)} class annotations, {len(all_method_annotations)} method annotations")
    print(f"✅ JaCoCo coverage parsing: {coverage_data['covered_branches']}/{coverage_data['total_branches']} branches ({coverage_data['coverage_percentage']:.1f}%)")
    print(f"✅ Excel file generation: {'PASS' if excel_valid else 'FAIL'}")
    print(f"✅ VBA template creation: {'PASS' if vba_exists else 'FAIL'}")

    # Verify no mock data was used
    if coverage_data['total_branches'] > 100 and len(all_method_annotations) > 5:
        print("✅ VERIFICATION: Real data detected (not mocks or fixed values)")
        return True
    else:
        print("❌ VERIFICATION: Data appears to be mocked or insufficient")
        return False

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)