#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_sheet_builder.py - Python版Excelシート生成

VBAのExcelSheetBuilder.basから移植されたExcel レポート生成機能
以下のシートを含む完全なテスト仕様書Excelファイルを生成:
1. Test Details - 完全なテストケース情報
2. Summary - 集計統計とメトリクス
3. Coverage - 詳細カバレッジ分析
4. Configuration - 処理設定とメタデータ

Created: 2026-01-07 (Pythonに移植)
Version: 2.0.0
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .data_types import (
    TestCaseInfo, CoverageInfo, SummaryStats, ExcelColors,
    ExcelFormatting, ConfigurationSettings, CoverageStatus,
    get_coverage_status, get_coverage_color, format_duration,
    ExcelGenerationError
)

class ExcelSheetBuilder:
    """Excelシート生成クラス"""

    def __init__(self, config: Optional[ConfigurationSettings] = None):
        """
        初期化

        Args:
            config: 設定情報（オプション）
        """
        self.config = config or ConfigurationSettings()
        self.logger = logging.getLogger(__name__)

        # Excel書式設定
        self.formatting = ExcelFormatting()

        # フォントとスタイル
        self._setup_styles()

        # 統計情報
        self.sheets_created = 0
        self.rows_written = 0

    def _setup_styles(self):
        """Excel書式スタイルをセットアップ"""
        # ヘッダーフォント
        self.header_font = Font(
            name='Arial',
            size=12,
            bold=True,
            color='FFFFFF'
        )

        # データフォント
        self.data_font = Font(
            name='Arial',
            size=10,
            bold=False
        )

        # ヘッダー背景色
        self.header_fill = PatternFill(
            start_color='4F81BD',  # ブルー
            end_color='4F81BD',
            fill_type='solid'
        )

        # 交互行背景色
        self.alt_row_fill = PatternFill(
            start_color='F2F2F2',  # ライトグレー
            end_color='F2F2F2',
            fill_type='solid'
        )

        # 境界線
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # セル配置
        self.center_alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

        self.left_alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )

    def generate_test_specification_report(self, output_file: str, test_cases: List[TestCaseInfo],
                                         coverage_data: List[CoverageInfo]) -> bool:
        """
        完全なテスト仕様書レポートを生成

        Args:
            output_file: 出力ファイルパス
            test_cases: テストケース情報のリスト
            coverage_data: カバレッジ情報のリスト

        Returns:
            bool: 生成成功の場合True
        """
        try:
            self.logger.info(f"Excelレポート生成開始: {output_file}")

            # 新しいワークブック作成
            wb = Workbook()

            # デフォルトシートを削除
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 統計情報計算
            summary_stats = self._calculate_summary_stats(test_cases, coverage_data)

            # 各シートを作成
            self._create_test_details_sheet(wb, test_cases)
            self._create_summary_sheet(wb, summary_stats)
            self._create_coverage_sheet(wb, coverage_data)
            self._create_configuration_sheet(wb)

            # ファイル保存
            wb.save(output_file)

            file_size = Path(output_file).stat().st_size
            self.logger.info(f"Excelレポート生成完了: {output_file} ({file_size:,}バイト)")

            return True

        except Exception as e:
            error_msg = f"Excelレポート生成エラー: {str(e)}"
            self.logger.error(error_msg)
            raise ExcelGenerationError(error_msg)

    def _create_test_details_sheet(self, wb: Workbook, test_cases: List[TestCaseInfo]):
        """
        テスト詳細シートを作成

        Args:
            wb: ワークブック
            test_cases: テストケース情報のリスト
        """
        ws = wb.create_sheet("Test Details")

        try:
            # ヘッダー行を作成
            headers = [
                'No.', 'Class Name', 'Method Name', 'Test Module', 'Test Case',
                'Test Overview', 'Test Purpose', 'Creator', 'Created Date',
                'Coverage %', 'Branches (Covered/Total)', 'Instructions (Covered/Total)',
                'Priority', 'Category', 'Baseline Version'
            ]

            self._write_header_row(ws, 1, headers)

            # データ行を作成
            for i, test_case in enumerate(test_cases, 2):
                row_data = [
                    i - 1,  # No.
                    test_case.class_name,
                    test_case.method_name,
                    test_case.test_module,
                    test_case.test_case,
                    test_case.test_overview,
                    test_case.test_purpose,
                    test_case.creator,
                    test_case.created_date.strftime('%Y-%m-%d') if test_case.created_date else '',
                    f"{test_case.coverage_percent:.1f}%",
                    f"{test_case.branches_covered}/{test_case.branches_total}",
                    f"{test_case.instructions_covered}/{test_case.instructions_total}",
                    test_case.priority,
                    test_case.test_category,
                    test_case.baseline_version
                ]

                self._write_data_row(ws, i, row_data, i % 2 == 0)

            # 列幅を調整
            column_widths = [5, 25, 25, 20, 20, 40, 40, 15, 12, 12, 20, 25, 12, 15, 15]
            self._adjust_column_widths(ws, column_widths)

            self.sheets_created += 1
            self.rows_written += len(test_cases)

        except Exception as e:
            self.logger.error(f"Test Detailsシート作成エラー: {str(e)}")
            raise

    def _create_summary_sheet(self, wb: Workbook, summary_stats: SummaryStats):
        """
        サマリーシートを作成

        Args:
            wb: ワークブック
            summary_stats: 統計情報
        """
        ws = wb.create_sheet("Summary")

        try:
            # タイトル
            ws['A1'] = 'Test Specification Summary Report'
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')

            # 生成日時
            ws['A3'] = 'Generated Date:'
            ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # 処理時間
            ws['A4'] = 'Processing Duration:'
            ws['B4'] = summary_stats.processing_duration

            # ファイル統計セクション
            row = 6
            ws[f'A{row}'] = 'File Statistics'
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill

            row += 1
            stats_data = [
                ('Total Java Files Processed', summary_stats.total_java_files),
                ('Total Test Classes Found', summary_stats.total_test_classes),
                ('Total Test Methods Found', summary_stats.total_test_methods),
                ('Total Coverage Reports', summary_stats.total_coverage_reports)
            ]

            for label, value in stats_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

            # カバレッジ統計セクション
            row += 1
            ws[f'A{row}'] = 'Coverage Statistics'
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill

            row += 1
            coverage_data = [
                ('Overall Branch Coverage', f"{summary_stats.overall_branch_coverage:.1f}%"),
                ('Total Branches Covered', f"{summary_stats.total_branches_covered}/{summary_stats.total_branches}"),
                ('Coverage Status', self._get_coverage_status_text(summary_stats.overall_branch_coverage))
            ]

            for label, value in coverage_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

            # エラー統計セクション
            if summary_stats.error_count > 0 or summary_stats.warning_count > 0:
                row += 1
                ws[f'A{row}'] = 'Processing Issues'
                ws[f'A{row}'].font = self.header_font
                ws[f'A{row}'].fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # イエロー

                row += 1
                ws[f'A{row}'] = 'Errors'
                ws[f'B{row}'] = summary_stats.error_count

                row += 1
                ws[f'A{row}'] = 'Warnings'
                ws[f'B{row}'] = summary_stats.warning_count

            # 列幅調整
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20

            self.sheets_created += 1

        except Exception as e:
            self.logger.error(f"Summaryシート作成エラー: {str(e)}")
            raise

    def _create_coverage_sheet(self, wb: Workbook, coverage_data: List[CoverageInfo]):
        """
        カバレッジシートを作成

        Args:
            wb: ワークブック
            coverage_data: カバレッジ情報のリスト
        """
        ws = wb.create_sheet("Coverage")

        try:
            # ヘッダー行を作成
            headers = [
                'Class Name', 'Method Name', 'Branch Coverage %',
                'Branches (Covered/Total)', 'Instructions (Covered/Total)',
                'Lines (Covered/Total)', 'Complexity (Covered/Total)', 'Status'
            ]

            self._write_header_row(ws, 1, headers)

            # データ行を作成
            for i, coverage in enumerate(coverage_data, 2):
                # カバレッジステータスを決定
                status = get_coverage_status(coverage.branch_coverage)
                status_color = get_coverage_color(status)

                row_data = [
                    coverage.class_name.split('.')[-1],  # クラス名のみ
                    coverage.method_name or 'Class Level',
                    f"{coverage.branch_coverage:.1f}%",
                    f"{coverage.branches_covered}/{coverage.branches_total}",
                    f"{coverage.instructions_covered}/{coverage.instructions_total}",
                    f"{coverage.lines_covered}/{coverage.lines_total}",
                    f"{coverage.complexity_covered}/{coverage.complexity_total}",
                    status.value
                ]

                self._write_data_row(ws, i, row_data, i % 2 == 0)

                # ステータス列に色を設定
                status_cell = ws.cell(i, 8)
                status_cell.fill = PatternFill(
                    start_color=f"{status_color:06X}",
                    end_color=f"{status_color:06X}",
                    fill_type='solid'
                )

            # 列幅を調整
            column_widths = [30, 25, 15, 20, 25, 20, 20, 12]
            self._adjust_column_widths(ws, column_widths)

            self.sheets_created += 1
            self.rows_written += len(coverage_data)

        except Exception as e:
            self.logger.error(f"Coverageシート作成エラー: {str(e)}")
            raise

    def _create_configuration_sheet(self, wb: Workbook):
        """
        設定シートを作成

        Args:
            wb: ワークブック
        """
        ws = wb.create_sheet("Configuration")

        try:
            # タイトル
            ws['A1'] = 'Processing Configuration and Metadata'
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:D1')

            # 処理設定セクション
            row = 3
            ws[f'A{row}'] = 'Processing Settings'
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill

            row += 1
            config_data = [
                ('Source Directory', self.config.source_directory),
                ('Output File Path', self.config.output_file_path),
                ('Include Subdirectories', 'Yes' if self.config.include_subdirectories else 'No'),
                ('Process Coverage Reports', 'Yes' if self.config.process_coverage_reports else 'No'),
                ('Include Test Files', 'Yes' if self.config.include_test_files else 'No'),
                ('Include IT Files', 'Yes' if self.config.include_it_files else 'No'),
                ('Exclude Abstract Classes', 'Yes' if self.config.exclude_abstract_classes else 'No'),
                ('Max File Size (MB)', f"{self.config.max_file_size / 1024 / 1024:.1f}"),
                ('Timeout (seconds)', str(self.config.timeout_seconds)),
                ('Log Detail Level', self.config.log_detail_level)
            ]

            for label, value in config_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

            # システム情報セクション
            row += 1
            ws[f'A{row}'] = 'System Information'
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill

            row += 1
            import platform
            system_data = [
                ('Python Version', platform.python_version()),
                ('Platform', platform.platform()),
                ('Architecture', platform.architecture()[0]),
                ('Processor', platform.processor() or 'Unknown'),
                ('Tool Version', '2.0.0 (Python版)')
            ]

            for label, value in system_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

            # 列幅調整
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40

            self.sheets_created += 1

        except Exception as e:
            self.logger.error(f"Configurationシート作成エラー: {str(e)}")
            raise

    def _write_header_row(self, ws: Worksheet, row: int, headers: List[str]):
        """
        ヘッダー行を書き込み

        Args:
            ws: ワークシート
            row: 行番号
            headers: ヘッダーテキストのリスト
        """
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment

    def _write_data_row(self, ws: Worksheet, row: int, data: List[Any], alt_row: bool = False):
        """
        データ行を書き込み

        Args:
            ws: ワークシート
            row: 行番号
            data: データのリスト
            alt_row: 交互行の場合True
        """
        for col, value in enumerate(data, 1):
            cell = ws.cell(row, col, value)
            cell.font = self.data_font
            cell.border = self.border
            cell.alignment = self.left_alignment if col > 2 else self.center_alignment

            if alt_row:
                cell.fill = self.alt_row_fill

    def _adjust_column_widths(self, ws: Worksheet, widths: List[int]):
        """
        列幅を調整

        Args:
            ws: ワークシート
            widths: 列幅のリスト
        """
        for col, width in enumerate(widths, 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

    def _calculate_summary_stats(self, test_cases: List[TestCaseInfo],
                                coverage_data: List[CoverageInfo]) -> SummaryStats:
        """
        統計情報を計算

        Args:
            test_cases: テストケース情報のリスト
            coverage_data: カバレッジ情報のリスト

        Returns:
            SummaryStats: 計算された統計情報
        """
        stats = SummaryStats()

        try:
            # ファイル統計
            unique_classes = set(tc.class_name for tc in test_cases)
            stats.total_java_files = len(unique_classes)
            stats.total_test_classes = len(unique_classes)
            stats.total_test_methods = len(test_cases)
            stats.total_coverage_reports = len(set(cd.report_file for cd in coverage_data))

            # カバレッジ統計
            total_branches = sum(tc.branches_total for tc in test_cases)
            covered_branches = sum(tc.branches_covered for tc in test_cases)

            if total_branches > 0:
                stats.overall_branch_coverage = (covered_branches / total_branches) * 100.0

            stats.total_branches = total_branches
            stats.total_branches_covered = covered_branches

            # 処理時間
            stats.processing_start_time = datetime.now()
            stats.processing_end_time = datetime.now()
            stats.processing_duration = format_duration(
                stats.processing_start_time,
                stats.processing_end_time
            )

        except Exception as e:
            self.logger.error(f"統計計算エラー: {str(e)}")

        return stats

    def _get_coverage_status_text(self, coverage_percent: float) -> str:
        """
        カバレッジ率からステータステキストを取得

        Args:
            coverage_percent: カバレッジ率

        Returns:
            str: ステータステキスト
        """
        status = get_coverage_status(coverage_percent)
        status_map = {
            CoverageStatus.EXCELLENT: "Excellent (90%+)",
            CoverageStatus.GOOD: "Good (80-89%)",
            CoverageStatus.FAIR: "Fair (60-79%)",
            CoverageStatus.POOR: "Poor (<60%)",
            CoverageStatus.UNKNOWN: "Unknown"
        }
        return status_map.get(status, "Unknown")

    def get_generation_stats(self) -> Dict[str, int]:
        """
        生成統計情報を取得

        Returns:
            Dict[str, int]: 統計情報
        """
        return {
            'sheets_created': self.sheets_created,
            'rows_written': self.rows_written
        }

# テスト関数

def test_excel_sheet_builder():
    """ExcelSheetBuilderのテスト"""
    print("🔍 ExcelSheetBuilderテスト開始...")

    builder = ExcelSheetBuilder()

    # サンプルデータ作成
    test_cases = [
        TestCaseInfo(
            class_name="BasicCalculatorTest",
            method_name="testConditionalCalculation",
            test_module="CalculatorModule",
            test_case="ConditionalAdditionTest",
            test_overview="Test addition with conditional branching",
            test_purpose="Ensure proper handling of different input types",
            creator="TestUser",
            coverage_percent=94.6,
            branches_covered=140,
            branches_total=148,
            instructions_covered=717,
            instructions_total=759
        )
    ]

    coverage_data = [
        CoverageInfo(
            class_name="BasicCalculatorTest",
            method_name="testConditionalCalculation",
            branch_coverage=94.6,
            branches_covered=140,
            branches_total=148,
            instructions_covered=717,
            instructions_total=759
        )
    ]

    # テスト用Excelファイル生成
    output_file = "/tmp/test_specification_python.xlsx"

    try:
        success = builder.generate_test_specification_report(output_file, test_cases, coverage_data)
        if success:
            print(f"✅ Excelレポート生成成功: {output_file}")

            # ファイルサイズ確認
            file_size = Path(output_file).stat().st_size
            print(f"   ファイルサイズ: {file_size:,}バイト")

            stats = builder.get_generation_stats()
            print(f"✅ 生成統計: シート{stats['sheets_created']}個, 行{stats['rows_written']}行")
        else:
            print("❌ Excelレポート生成失敗")

    except Exception as e:
        print(f"❌ テストエラー: {str(e)}")

    print("🎉 ExcelSheetBuilderテスト完了!")

if __name__ == "__main__":
    # ログ設定
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    test_excel_sheet_builder()